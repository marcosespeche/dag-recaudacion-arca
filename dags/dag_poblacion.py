from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.http.operators.http import HttpOperator
from airflow.providers.http.hooks.http import HttpHook
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
import os
import json
import time
import logging
from pathlib import Path
import xlrd
from openpyxl import load_workbook
logging.getLogger("airflow.hooks.base").setLevel(logging.ERROR)

PATH_ARCHIVOS_POBLACION = "/tmp/poblacion/"

default_args = {
    'owner': 'equipo_13',
    'start_date': datetime.today() - timedelta(days=1),
    'retries': 1,
    'retry_delay': timedelta(minutes=2),
    'email_on_failure': False,
    'depends_on_past': False,
}

# Función que trae una hoja de un archivo Excel (.xls)
def obtener_hoja_xls(hook, endpoint, nombre_hoja):
    res = hook.run(endpoint)
    data = BytesIO(res.content)

    workbook = xlrd.open_workbook(file_contents=res.content)
    sheet = workbook.sheet_by_name(nombre_hoja)

    data = []
    for row_idx in range(sheet.nrows):
        row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            cell_value = cell.value
            
            # El casteo a datetime no se realiza de forma manual
            if cell.ctype == xlrd.XL_CELL_DATE:
                date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                cell_value = datetime(*date_tuple)
            
            row.append(cell_value)
        data.append(row)    
    
    return data

# Función que trae una hoja de un archivo Excel (.xlsx)
def obtener_hoja_xlsx(hook, endpoint, nombre_hoja):
   res = hook.run(endpoint)
   data = BytesIO(res.content)

   workbook = load_workbook(data,data_only=True)
   sheet = workbook[nombre_hoja]

   data = []
   for row in sheet.iter_rows(values_only=True):
       data.append(list(row))
   
   return data

class CSVExporter:
    def __init__(self):
        self.values = []
    
    def add(self, provincia, valor):
        
        self.values.append({
            "provincia": provincia,
            "valor": valor
        })

    def export(self, path, name):
        df = pd.DataFrame(self.values)
        if not df.empty:
            df = df.sort_values(['provincia', 'valor'])
        
        Path(path).mkdir(parents=True, exist_ok=True)
        ruta = f"{path}{name}.csv"
        df.to_csv(ruta, index=False)

        return ruta


def descargar_archivos_poblacion(**kwards):
    hook = HttpHook(http_conn_id='indec', method='GET')
    try:
        data = obtener_hoja_xls(hook, "ftp/censos/2010/CuadrosDefinitivos/Total_pais/P1-P_Total_pais.xls", "totalpaís")
        
        csv_exporter = CSVExporter()
        for index, row in enumerate(data):
            if index > 5 and index < 34 and row[0]!='':
                provincia=row[0]
                poblacion=row[2]
                csv_exporter.add(provincia, poblacion)

        return csv_exporter.export(PATH_ARCHIVOS_POBLACION, "poblacion_2010")

    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar datos de población Censo 2010: {e}")
        raise e

with DAG(
    dag_id='poblacion_2010',
    description='DAG ETL que recopila datos de la poblacion en Argentina 2010',
    default_args=default_args,
    schedule=None,
    catchup=False,
    tags=['poblacion', 'Argentina', 'etl']
) as dag:

    task_descargar_archivos_poblacion = PythonOperator(
        task_id = "task_descargar_archivos_poblacion",
        python_callable = descargar_archivos_poblacion
    )


    task_descargar_archivos_poblacion

