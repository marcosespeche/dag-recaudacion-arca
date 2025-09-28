from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.http.operators.http import HttpOperator
from airflow.providers.http.hooks.http import HttpHook
from datetime import date, datetime, timedelta
import pandas as pd
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
import logging
import xlrd
from openpyxl import load_workbook
logging.getLogger("airflow.hooks.base").setLevel(logging.ERROR)

INITIAL_YEAR = 2008
LAST_YEAR = 2025
URL_ARCHIVO_IPC_ARGENTINA = "/ftp/cuadros/economia/sh_ipc_08_25.xls"
URL_ARCHIVO_IPC_CORDOBA = "/dataset/fedc5285-5517-41aa-9095-bb62c6dbc485/resource/2b4a7c60-1c8a-45b1-be8f-2bd59bfe2364/download/ipc-cba-julio.xlsx"
URL_ARCHIVO_RECAUDACION = "/institucional/estudios/archivos/serie"
URL_ARCHIVO_IPC_MENDOZA = "/data/download-files/?ids=1365"
URL_ARCHIVO_IPC_GBA = "/catalog/sspm/dataset/96/distribution/96.3/download/indice-precios-al-consumidor-bienes-servicios-base-2008-mensual.csv"
URL_ARCHIVO_BALANZA_COMERCIAL = "/ftp/cuadros/economia/balanmensual.xls"
URL_ARCHIVO_DOLAR_BLUE = '/historico-dolar-blue/custom/1-1-2008_26-9-2025'
URL_API_MONEDAS_PUBLICO = '/Monetarias/25'
URL_API_DISTRIBUCION_INGRESOS = '/indicadores-distribucion-ingreso-ocupacion-principal-e-ingreso-per-capita-familiar.csv'

PATH_ARCHIVOS_RECAUDACION = "/tmp/recaudacion/"
PATH_ARCHIVOS_IPC = "/tmp/ipc/"
PATH_ARCHIVOS_EMAE = "/tmp/emae/"
PATH_ARCHIVOS_OUTPUT = "/tmp/result/"
PATH_ARCHIVOS_BALANZA_COMERCIAL = "/tmp/balanza-comercial/"
PATH_ARCHIVOS_DOLAR_BLUE = "/tmp/dolar-blue/"
PATH_ARCHIVOS_MONEDAS_PUBLICO = "/tmp/monedas-publico/"
PATH_ARCHIVOS_DISTRIBUCION_INGRESOS = "/tmp/distribucion-ingresos/"

LIST_TASKS_ID = ['recaudacion', 'emae', 'ipc_argentina', 'ipc_cordoba', 'ipc_tucuman', 'ipc_santafe', 'ipc_gba', 'ipc_mendoza', 'balanza_comercial', 'dolar_blue', 'monedas_publico', 'ingresos_familiares']

default_args = {
    'owner': 'equipo_13',
    'start_date': datetime.today() - timedelta(days=1),
    'retries': 1,
    'retry_delay': timedelta(minutes=2),
    'email_on_failure': False,
    'depends_on_past': False,
}

#Map para convertir nombre de mes a número
MESES_MAP = {
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12
}

# Función que trae una hoja de un archivo Excel (.xls)
def obtener_hoja_xls(hook, endpoint, nombre_hoja):
    res = hook.run(endpoint)
    data = BytesIO(res.content)

    workbook = xlrd.open_workbook(file_contents=res.content)
    sheet = workbook.sheet_by_name(nombre_hoja)

    data = []
    for row_idx in range(sheet.nrows):
        row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            cell_value = cell.value
            
            # El casteo a datetime no se realiza de forma manual
            if cell.ctype == xlrd.XL_CELL_DATE:
                date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                cell_value = datetime(*date_tuple)
            
            row.append(cell_value)
        data.append(row)    
    
    return data

# Función que trae una hoja de un archivo Excel (.xlsx)
def obtener_hoja_xlsx(hook, endpoint, nombre_hoja):
   res = hook.run(endpoint)
   data = BytesIO(res.content)

   workbook = load_workbook(data,data_only=True)
   sheet = workbook[nombre_hoja]

   data = []
   for row in sheet.iter_rows(values_only=True):
       data.append(list(row))
   
   return data

# Función que obtiene un archivo CSV
def obtener_csv(hook, endpoint):
    res = hook.run(endpoint)
    data = BytesIO(res.content)
    df = pd.read_csv(data, header=None)
    return df.values.tolist()

class CSVExporter:
    def __init__(self):
        self.values = []
        self.overwrite_repeated = False

    def set_overwrite_repeated_policy(self, overwrite : bool):
        self.overwrite_repeated = overwrite
    
    def add(self, anio, mes, valor):
        if not isinstance(anio, int) or anio < 1900 or anio > 2025:
            raise Exception(f"El año {anio} no es un número entero o no es válido")
        if not isinstance(mes, int) or mes < 1 or mes > 12:
            raise Exception(f"El mes {mes} no es un número entero o no es válido")
        if not isinstance(valor, (int, float)):
            raise Exception(f"El valor {valor} no es válido")
        
        if not any(v['anio'] == anio and v['mes'] == mes for v in self.values):
            self.values.append({
                "anio": anio,
                "mes": mes,
                "valor": valor
            })
            return
        
        if self.overwrite_repeated:
            for v in self.values:
                if v['anio'] == anio and v['mes'] == mes:
                    v['valor'] = valor
                    break
            

    def export(self, path, name):
        df = pd.DataFrame(self.values)
        if not df.empty:
            df = df.sort_values(['anio', 'mes'])
        
        Path(path).mkdir(parents=True, exist_ok=True)
        ruta = f"{path}{name}.csv"
        df.to_csv(ruta, index=False)

        return ruta


def descargar_ipc_argentina(**kwards):
    hook = HttpHook(http_conn_id='indec', method='GET')
    
    try:
        source = obtener_hoja_xls(hook, URL_ARCHIVO_IPC_ARGENTINA, "Variación mensual IPC Nacional")

        # Encontrar filas de periodo y de nivel general
        nombre_fila_periodo = "Total nacional"
        ix_fila_periodo = None
        nombre_fila_nivel_general = "Nivel general"
        ix_fila_nivel_general = None
        for index, row in enumerate(source):
            if row[0] == nombre_fila_periodo:
                ix_fila_periodo = index
                continue
            if row[0] == nombre_fila_nivel_general:
                ix_fila_nivel_general = index
            if ix_fila_periodo is not None and ix_fila_nivel_general is not None:
                break
        if (ix_fila_periodo is None):
            raise Exception("No se encontró la fila con los periodos")
        if (ix_fila_nivel_general is None):
            raise Exception("No se encontró la fila con los valores del nivel general")

        # Encontrar valores del IPC
        exporter = CSVExporter()
        for index, column in enumerate(zip(*source)):
            periodo = column[ix_fila_periodo]
           
            if isinstance(periodo, datetime):
                anio = periodo.year
                mes = periodo.month
            elif isinstance(periodo, (int, float)) and periodo > 0:
                if periodo >= 60:
                    date_obj = datetime(1899, 12, 30) + timedelta(days=periodo)
                else:
                    date_obj = datetime(1899, 12, 31) + timedelta(days=periodo)
                anio = date_obj.year
                mes = date_obj.month
            else:
                continue
            
            valor = float(column[ix_fila_nivel_general]) / 100
            exporter.add(anio, mes, valor)

        # Convertir a CSV
        ruta = exporter.export(PATH_ARCHIVOS_IPC, "argentina")

        return ruta

    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar IPC de Argentina: {e}")
        raise e

def descargar_ipc_cordoba(**kwards):
    hook = HttpHook(http_conn_id='estadisticacordoba', method='GET')

    try:
        source = obtener_hoja_xlsx(hook, URL_ARCHIVO_IPC_CORDOBA, "Variaciones Mensuales")
        # Encontrar filas de periodo y de nivel general
        nombre_fila_periodo = "Descripción"
        ix_fila_periodo = None
        nombre_fila_nivel_general = "NIVEL GENERAL"
        ix_fila_nivel_general = None
        for index, row in enumerate(source):
            if row[2] == nombre_fila_periodo:
                ix_fila_periodo = index
                continue
            if row[2] == nombre_fila_nivel_general:
                ix_fila_nivel_general = index
            if ix_fila_periodo is not None and ix_fila_nivel_general is not None:
                break
        if (ix_fila_periodo is None):
            raise Exception("No se encontró la fila con los periodos")
        if (ix_fila_nivel_general is None):
            raise Exception("No se encontró la fila con los valores del nivel general")

        # Encontrar valores del IPC
        exporter = CSVExporter()
        for index, column in enumerate(zip(*source)):
            periodo = column[ix_fila_periodo]
                        
            if not isinstance(periodo, datetime):
                continue

            anio = periodo.year
            mes = periodo.month
            valor = float(column[ix_fila_nivel_general])
            
            exporter.add(anio, mes, valor)

        # Convertir a CSV
        ruta = exporter.export(PATH_ARCHIVOS_IPC, "cordoba")

        return ruta

    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar IPC de Córdoba: {e}")
        raise e

def descargar_recaudacion(**kwards):
    hook = HttpHook(http_conn_id='recaudacion-arca', method='GET')
    csv_exporter = CSVExporter()
    try:
        for year in range(INITIAL_YEAR, LAST_YEAR + 1):
            data = obtener_hoja_xls(hook, f"{URL_ARCHIVO_RECAUDACION}{year}.xls", f"{year}")

            nombre_fila_recaudacion = '  TOTAL GENERAL'
            idx_fila_recaudacion = None

            for index, row in enumerate(data):
                if (row[1] != nombre_fila_recaudacion):
                    continue
                idx_fila_recaudacion = index

                for month in range(1, 13):
                    column_index = month + 1
                    recaudacion = row[column_index]
                    if recaudacion is not None and recaudacion != '':
                        # No agarra el valor total (la sumatoria de todos los meses) porque está definido como una función de Excel, no es un valor numerico en sí
                        csv_exporter.add(year, month, recaudacion)
                break

            if idx_fila_recaudacion is None:
                logging.warning(f"No se encontró la fila con los totales de recaudación para el año {year}")

        return csv_exporter.export(PATH_ARCHIVOS_RECAUDACION, 'recaudacion-arca')
    
    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar la recaudación de ARCA: {e}")
        raise e

def descargar_emae(**kwards):
    hook = HttpHook(http_conn_id='emae', method='GET')
    csv_exporter = CSVExporter()
    
    MESES_MAP = { 'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4, 'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8, 'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12 }

    current_year = None
    meses_por_anio = {}

    try:
        data = obtener_hoja_xls(hook, 'sh_emae_mensual_base2004.xls', 'EMAE')

        """
            Formato del Excel:

            2024 Enero valor1
                 Febrero valor 2
                 marzo valor3
            ...
            2025 Enero valor4

            Es decir, hay filas que contienen el año y filas que no 
            """

        for index, row in enumerate(data):
            anio = row[0]
            mes = row[1] 
            valor = row[2]

            # Si aun no se llega a un año en curso, se saltea la fila
            if current_year is None and not (isinstance(anio, (int, float))):
                continue

            # Si se detecta un nuevo año, se pisa el valor de current_year anterior y añaden los datos registrados hasta el momento en el CSV
            if isinstance(anio, (int, float)):

                if current_year and meses_por_anio:
                    for m, v in meses_por_anio.items():
                        csv_exporter.add(current_year, m, v)
                    meses_por_anio.clear()

                current_year = int(anio)

            # Registrar valor del mes si es válido
            if mes and mes in MESES_MAP and valor is not None:
                mes_numero = MESES_MAP[mes]
                meses_por_anio[mes_numero] = valor

        # Volcar lo último acumulado
        if current_year and meses_por_anio:
            for m, v in meses_por_anio.items():
                csv_exporter.add(current_year, m, v)

    except Exception as e:
        logging.error(f"Error al descargar el EMAE: {e}")
        raise e

    return csv_exporter.export(PATH_ARCHIVOS_EMAE, 'emae')

def descargar_ipc_tucuman(**kwargs):
    hook = HttpHook(http_conn_id = 'ipc-tucuman', method = 'GET')
    csv_exporter = CSVExporter()

    try:
        data = obtener_hoja_xlsx(hook, '/archivos/7Precios/1Precios/SERIE_IPCT_JUL_25.xlsx', 'SERIE DE TIEMPO')

        for index, row in enumerate(data):

            # Si no encuentra una fecha, pasa a la siguiente fila
            if (not isinstance(row[0], (datetime, date))):
                continue
            
            fecha = row[0]
            nro_mes = fecha.month
            nro_anio = fecha.year
            ipc = row[2]
            
            if (ipc is not None):
                csv_exporter.add(nro_anio, nro_mes, ipc)

        return csv_exporter.export(PATH_ARCHIVOS_IPC, 'tucuman')
    
    except Exception as e:
        logging.error(f"Error al descargar el IPC de Tucumán: {e}")
        raise e
    
def descargar_ipc_santafe(**kwargs):
    hook = HttpHook(http_conn_id = 'ipc-santafe', method = 'GET')
    csv_exporter = CSVExporter()

    try:
        data = obtener_hoja_xls(hook, '/wp-content/uploads/sites/24/2025/06/IPC_Serie-IPC-por-capitulos-0725.xls', r'Var % capítulos')

        for index, row in enumerate(data):

            fecha = row[0]

            if (not isinstance(fecha, (datetime, date))):
                continue

            nro_mes = fecha.month
            nro_anio = fecha.year
            ipc = row[1]
            csv_exporter.add(nro_anio, nro_mes, ipc)

        return csv_exporter.export(PATH_ARCHIVOS_IPC, 'santafe')

    except Exception as e:
        logging.error(f'Error al descargar el IPC de Santa Fé: {e}')
        raise e

def descargar_ipc_mendoza(**kwargs):
    hook = HttpHook(http_conn_id='ipc-mendoza', method='GET')
    try:
        data = obtener_hoja_xls(hook, URL_ARCHIVO_IPC_MENDOZA, "IPC MZA Base 2010")
        anio = None
        csv_exporter = CSVExporter()
        for index, row in enumerate(data):
            #filas con los datos en el formato que necesitamos
            if 5 < index < 581:
                if row and len(row) >= 4:
                    if row[1] not in (None, ''):
                        anio = int(float(row[1]))

                    mes_str = str(row[2]).strip().lower()
                    if mes_str in MESES_MAP:
                        mes = MESES_MAP[mes_str]
                    else:
                        continue

                    valor = float(str(row[3]).replace(',', '.'))

                    if anio and mes and valor:
                        csv_exporter.add(anio, mes, valor)              

        return csv_exporter.export(PATH_ARCHIVOS_IPC, "mendoza")

    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar datos de ipc Mendoza: {e}")
        raise e

def descargar_ipc_gba(**kwards):
    hook = HttpHook(http_conn_id='ipc-gba', method='GET')
    csv_exporter = CSVExporter()
    data = obtener_csv(hook, URL_ARCHIVO_IPC_GBA)
    try:
        for row in data[1:]:
            fecha = row[0]
            valor = float(row[1])
            fecha_dt = pd.to_datetime(fecha)
            anio = fecha_dt.year
            mes = fecha_dt.month
            csv_exporter.add(anio, mes, valor)
        return csv_exporter.export(PATH_ARCHIVOS_IPC, "gba")
    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar datos de ipc GBA: {e}")
        raise e
    
def descargar_balanza_comercial(**kwargs):
    hook = HttpHook(http_conn_id='balanza-comercial', method='GET')
    csv_importaciones = CSVExporter()
    csv_exportaciones = CSVExporter()

    data = obtener_hoja_xls(hook, URL_ARCHIVO_BALANZA_COMERCIAL, 'FOB-CIF')

    current_year = None
    exportaciones_por_mes = {}
    importaciones_por_mes = {}

    columna_exportaciones = 2
    columna_importaciones = 7

    try:
        for index, row in enumerate(data):

            anio = limpiar_anio(row[0])
            mes = str(row[1]).strip().lower()
            exportacion = row[columna_exportaciones]
            importacion = row[columna_importaciones]

            # Si aun no se llega a un año en curso, se saltea la fila
            if current_year is None and not (isinstance(anio, (int, float))):
                continue

            # Si se detecta un nuevo año, se pisa el valor de current_year anterior y añaden los datos registrados hasta el momento en el CSV
            if isinstance(anio, (int, float)):

                if current_year and exportaciones_por_mes and importaciones_por_mes:
                    for m, v in exportaciones_por_mes.items():
                        csv_exportaciones.add(current_year, m, v)
                    exportaciones_por_mes.clear()
                    for m, v in importaciones_por_mes.items():
                        csv_importaciones.add(current_year, m, v)
                    importaciones_por_mes.clear()

                current_year = int(anio)

            # Registrar valor del mes si es válido
            if mes and mes in MESES_MAP and exportacion is not None and importacion is not None:
                mes_numero = MESES_MAP[mes]
                exportaciones_por_mes[mes_numero] = exportacion
                importaciones_por_mes[mes_numero] = importacion

        # Volcar lo último acumulado
        if current_year and exportaciones_por_mes and importaciones_por_mes:
            for m, v in exportaciones_por_mes.items():
                csv_exportaciones.add(current_year, m, v)
            for m, v in importaciones_por_mes.items():
                csv_importaciones.add(current_year, m, v)

    except Exception as e:
        logging.error(f'Error al descargar datos de la Balanza Comercial: {e}')
        raise e
    
    return csv_exportaciones.export(PATH_ARCHIVOS_BALANZA_COMERCIAL, 'exportaciones'), csv_importaciones.export(PATH_ARCHIVOS_BALANZA_COMERCIAL, 'importaciones')

def limpiar_anio(valor):
    if valor is None:
        return None
    
    valor_str = str(valor).strip()

    if valor_str.startswith("'"):
        valor_str = valor_str[1:]

    if valor_str.endswith("*"):
        valor_str = valor_str[:-1]

    if valor_str.isdigit():
        return int(valor_str)
    
    return None

def descargar_dolar_blue(**kwargs):
    try:
        csv_exporter = CSVExporter()

        hook = HttpHook(http_conn_id='dolar-blue', method='GET')

        response = hook.run(endpoint=URL_ARCHIVO_DOLAR_BLUE)
  
        if response.status_code != 200:
            raise Exception(f"Error HTTP {response.status_code}: {response.text}")

        data = response.json()

        ultimo_anio = None
        ultimo_mes = None
        ultimo_valor = None

        for index, item in enumerate(data):
            fecha_str = item['x']
            valor = item['y']

            anio, mes = parsear_fecha(fecha_str)

            if (ultimo_anio is not None and ultimo_mes is not None and 
                (anio != ultimo_anio or mes != ultimo_mes)):
                csv_exporter.add(ultimo_anio, ultimo_mes, ultimo_valor)

            ultimo_anio = anio
            ultimo_mes = mes
            ultimo_valor = valor

        if ultimo_anio is not None and ultimo_mes is not None and ultimo_valor is not None:
            csv_exporter.add(ultimo_anio, ultimo_mes, ultimo_valor)

        return csv_exporter.export(PATH_ARCHIVOS_DOLAR_BLUE, 'dolar_blue')
        
    except Exception as e:
        logging.error(f"Error al descargar datos del dólar blue: {e}")
        raise e


def parsear_fecha(fecha_str):
    try:
        fecha_limpia = fecha_str.split(" GMT")[0]

        fecha_obj = datetime.strptime(fecha_limpia, "%a %b %d %Y %H:%M:%S")
        
        return fecha_obj.year, fecha_obj.month
        
    except Exception as e:
        try:
            import re
            year_match = re.search(r'\b(20[0-3]\d)\b', fecha_str)
            meses_en = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
            }
            
            if year_match:
                anio = int(year_match.group(1))
                fecha_lower = fecha_str.lower()
                for mes_nombre, mes_num in meses_en.items():
                    if mes_nombre in fecha_lower:
                        return anio, mes_num
                 
        except:
            pass
            
        return None, None
    

def descargar_monedas_publico(**kwargs):
    try:
        hook = HttpHook(http_conn_id='bcra-estadisticas', method='GET')
        csv_exporter = CSVExporter()

        desde = "2008-01-01"
        hasta = datetime.today().strftime('%Y-%m-%d')

        isLastPage = False
        offset = 0

        while (not isLastPage):
            response = hook.run(endpoint=f'{URL_API_MONEDAS_PUBLICO}?desde={desde}&hasta={hasta}&limit=3000&offset={offset}', extra_options={"verify": False})
        
            if response.status_code != 200:
                raise Exception(f"Error HTTP {response.status_code}: {response.text}")

            data = response.json()

            if (data["metadata"]["resultset"]["offset"] + data["metadata"]["resultset"]["limit"] > data["metadata"]["resultset"]["count"]):
                isLastPage = True
            offset += 3000

            detalles = data["results"][0]["detalle"]

            for d in detalles:
                fecha = datetime.strptime(d["fecha"], '%Y-%m-%d').date()
                anio = fecha.year
                mes = fecha.month

                csv_exporter.add(anio, mes, d["valor"])

        return csv_exporter.export(PATH_ARCHIVOS_MONEDAS_PUBLICO, 'monedas_publico')

    except Exception as e:
        logging.error(f"Error al descargar datos de billetes y monedas en poder del público: {e}")
        raise e

def descargar_ingresos_familiares(**kwargs):
    hook = HttpHook(http_conn_id='ingresos-familiares', method='GET')
    csv_exporter = CSVExporter()
    data = obtener_csv(hook, URL_API_DISTRIBUCION_INGRESOS)
    try:
        for row in data[1:]:
            fecha = row[0]
            valor = float(row[2])
            fecha_dt = pd.to_datetime(fecha)
            anio = fecha_dt.year
            mes = fecha_dt.month
            csv_exporter.add(anio, mes, valor)
        return csv_exporter.export(PATH_ARCHIVOS_DISTRIBUCION_INGRESOS, "ingresos_familiares")
    except Exception as e:
        logging.error(f"[ERROR] Problema al descargar datos de distribución de ingresoo familiares: {e}")
        raise e

def merge (**kwargs):
    data = []
    
    for task in LIST_TASKS_ID:
        if task == 'balanza_comercial':
            # Porque la balanza retorna una tupla de rutas (una para exportaciones, y otra para importaciones)
            rutas = kwargs['ti'].xcom_pull(task_ids="task_descargar_" + task)
            ruta_exportaciones, ruta_importaciones = rutas
            
            # Agregar ambos archivos por separado
            data.append({
                'nombre': 'exportaciones', 
                'df': pd.read_csv(ruta_exportaciones)
            })
            data.append({
                'nombre': 'importaciones', 
                'df': pd.read_csv(ruta_importaciones)
            })
        else:
            data.append({
                'nombre': task, 
                'df': pd.read_csv(kwargs['ti'].xcom_pull(task_ids="task_descargar_" + task))
            })

    result_dfs = []

    for item in data:
        df = item['df'].set_index(['anio', 'mes'])
        df = df.rename(columns= {'valor': item['nombre']})
        result_dfs.append(df)

    df = pd.concat(result_dfs, axis=1, join='outer')

    Path(PATH_ARCHIVOS_OUTPUT).mkdir(parents=True, exist_ok=True)

    df = df.sort_values(by=['anio', 'mes'])

    df.to_csv(PATH_ARCHIVOS_OUTPUT + 'montos.csv')


with DAG(
    dag_id='montos',
    description='DAG ETL que recopila datos económicos en Argentina y de la recaudación realizada por ARCA desde 2008 hasta 2025',
    default_args=default_args,
    schedule=None,
    catchup=False,
    tags=['arca', 'recaudacion', 'etl']
) as dag:
    
    task_descargar_ipc_argentina = PythonOperator(
        task_id = "task_descargar_ipc_argentina",
        python_callable = descargar_ipc_argentina
    )

    task_descargar_ipc_cordoba = PythonOperator(
        task_id = "task_descargar_ipc_cordoba",
        python_callable = descargar_ipc_cordoba
    )

    task_descargar_archivos_recaudacion = PythonOperator(
        task_id = "task_descargar_recaudacion",
        python_callable = descargar_recaudacion
    )

    task_descargar_emae = PythonOperator(
        task_id = 'task_descargar_emae',
        python_callable = descargar_emae
    )

    task_descargar_ipc_tucuman = PythonOperator(
        task_id = 'task_descargar_ipc_tucuman',
        python_callable = descargar_ipc_tucuman
    )

    task_descargar_ipc_santafe = PythonOperator(
        task_id = 'task_descargar_ipc_santafe',
        python_callable = descargar_ipc_santafe
    )

    task_descargar_ipc_mendoza = PythonOperator(
        task_id = 'task_descargar_ipc_mendoza',
        python_callable = descargar_ipc_mendoza
    )

    task_descargar_ipc_gba = PythonOperator(
        task_id = 'task_descargar_ipc_gba',
        python_callable = descargar_ipc_gba
    )

    task_descargar_balanza_comercial = PythonOperator(
        task_id = 'task_descargar_balanza_comercial',
        python_callable = descargar_balanza_comercial
    )

    task_descargar_dolar_blue = PythonOperator(
        task_id = 'task_descargar_dolar_blue',
        python_callable = descargar_dolar_blue
    )

    task_descargar_monedas_publico = PythonOperator(
        task_id = 'task_descargar_monedas_publico',
        python_callable = descargar_monedas_publico
    )

    task_descargar_ingresos_familiares = PythonOperator(
        task_id = 'task_descargar_ingresos_familiares',
        python_callable = descargar_ingresos_familiares
    )

    task_merge = PythonOperator(
        task_id = 'task_merge',
        python_callable = merge
    )

    [task_descargar_ipc_argentina, task_descargar_ipc_cordoba, task_descargar_ipc_tucuman, task_descargar_archivos_recaudacion, task_descargar_emae, task_descargar_ipc_mendoza, task_descargar_ipc_santafe, task_descargar_ipc_gba, task_descargar_balanza_comercial, task_descargar_dolar_blue, task_descargar_monedas_publico] >> task_merge

# El csv para la poblacion zona (String), poblacion, solo para las provincias que tengamos IPC

# En el IPC de GBA hay que sumar dos valores distintos porque salen desagregados
# Los csv para los IPC contienen las columnas anio, mes (1 al 12), valor (completo, sin redondear ni truncar), solo entre enero 2008 hasta julio 2025
# Los csv para el merge final contienen: anio, mes, recaudacion, ipc, emae